"""
Bulk Upload router — supports CSV and XLSX imports for master data and transactions.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from typing import Optional
import csv
import io
import json
import logging
from datetime import datetime, date

from ..database import get_db
from ..models.user import User
from ..models.supplier import Supplier
from ..models.category import Category
from ..models.product import Product
from ..models.warehouse import Warehouse
from ..models.store import Store
from ..models.bank_account import BankAccount
from ..models.receiving import Receiving
from ..models.supplier_return import SupplierReturn
from ..models.bulk_import import BulkImportHistory, BulkImportError
from ..schemas.bulk_import import BulkValidateResponse, BulkImportResponse, BulkImportErrorSchema
from ..services.auth import get_current_user
from ..services.permissions import require_permission
from ..services.inventory_service import update_inventory_balance
from ..constants import INVENTORY_TYPES, DEFAULT_INVENTORY_TYPE

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/bulk-upload", tags=["Bulk Upload"])

# ── Template definitions ────────────────────────────────────────────────────
TEMPLATES = {
    "suppliers": {
        "columns": ["supplier_name", "supplier_contact", "supplier_email", "supplier_address"],
        "required": ["supplier_name"],
        "example": [{"supplier_name": "PT Example Supplier", "supplier_contact": "08123456789", "supplier_email": "supplier@example.com", "supplier_address": "Jl. Example No.1, Bali"}],
    },
    "categories": {
        "columns": ["category_name", "description"],
        "required": ["category_name"],
        "example": [{"category_name": "Water Filter", "description": "Water purification products"}],
    },
    "products": {
        "columns": ["product_name", "category_name", "sale_price", "sku", "barcode", "unit", "status", "minimum_stock_level", "product_description"],
        "required": ["product_name", "sale_price"],
        "example": [{"product_name": "Nazava Filter S", "category_name": "Water Filter", "sale_price": "500000", "sku": "NF-S001", "barcode": "", "unit": "PCS", "status": "Active", "minimum_stock_level": "10", "product_description": "Portable water filter"}],
    },
    "bank_accounts": {
        "columns": ["bank_name", "account_number", "beneficiary_name", "is_active"],
        "required": ["bank_name", "account_number", "beneficiary_name"],
        "example": [{"bank_name": "BCA", "account_number": "1234567890", "beneficiary_name": "PT Kopernik", "is_active": "true"}],
    },
    "receivings": {
        "columns": [
            "received_date", "supplier_name", "product_name", "category", "warehouse_name",
            "quantity_received", "quantity_rejected", "purchase_price", "inventory_type",
        ],
        "required": ["received_date", "supplier_name", "product_name", "quantity_received", "purchase_price"],
        "example": [{
            "received_date": "2026-07-01", "supplier_name": "Nazava", "product_name": "Nazava Filter S",
            "category": "Water Filter", "warehouse_name": "Ubud Warehouse", "quantity_received": "100",
            "quantity_rejected": "0", "purchase_price": "250000", "inventory_type": "TKU Product",
        }],
    },
}

SUPPORTED_TYPES = list(TEMPLATES.keys())


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=422, detail="XLSX support requires openpyxl. Upload a CSV file instead.")
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append({headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))})
    return result


def _parse_file(filename: str, content: bytes) -> list[dict]:
    if filename.lower().endswith(".xlsx"):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _validate_suppliers(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    valid, errors = [], []
    tmpl = TEMPLATES["suppliers"]
    for i, row in enumerate(rows, 1):
        errs = []
        for col in tmpl["required"]:
            if not row.get(col, "").strip():
                errs.append(f"'{col}' is required")
        if errs:
            errors.append({"row_number": i, "error_message": "; ".join(errs), "raw_data": json.dumps(row)})
        else:
            valid.append({"row": i, "data": row})
    return valid, errors


def _validate_categories(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    valid, errors = [], []
    for i, row in enumerate(rows, 1):
        if not row.get("category_name", "").strip():
            errors.append({"row_number": i, "error_message": "'category_name' is required", "raw_data": json.dumps(row)})
        else:
            valid.append({"row": i, "data": row})
    return valid, errors


def _validate_products(rows: list[dict], db: Session) -> tuple[list[dict], list[dict]]:
    valid, errors = [], []
    # Pre-fetch categories by name for lookup
    cats = {c.category_name: c.category_id for c in db.query(Category).filter(Category.deleted_at.is_(None)).all()}
    for i, row in enumerate(rows, 1):
        errs = []
        if not row.get("product_name", "").strip():
            errs.append("'product_name' is required")
        price_str = row.get("sale_price", "").strip()
        if not price_str:
            errs.append("'sale_price' is required")
        else:
            try:
                float(price_str)
            except ValueError:
                errs.append(f"'sale_price' must be a number, got '{price_str}'")
        cat_name = row.get("category_name", "").strip()
        if cat_name and cat_name not in cats:
            errs.append(f"Category '{cat_name}' not found — create it first")
        if errs:
            errors.append({"row_number": i, "error_message": "; ".join(errs), "raw_data": json.dumps(row)})
        else:
            enriched = dict(row)
            enriched["_category_id"] = cats.get(cat_name)
            valid.append({"row": i, "data": enriched})
    return valid, errors


def _validate_receivings(rows: list[dict], db: Session) -> tuple[list[dict], list[dict]]:
    valid, errors = [], []
    suppliers = {s.supplier_name: s.supplier_id for s in db.query(Supplier).filter(Supplier.deleted_at.is_(None)).all()}
    warehouses = {w.warehouse_name: w.warehouse_id for w in db.query(Warehouse).filter(Warehouse.deleted_at.is_(None)).all()}
    products = db.query(Product).filter(Product.deleted_at.is_(None)).all()
    products_by_name = {}
    for p in products:
        products_by_name.setdefault(p.product_name, []).append(p)

    for i, row in enumerate(rows, 1):
        errs = []
        received_date_str = row.get("received_date", "").strip()
        supplier_name = row.get("supplier_name", "").strip()
        product_name = row.get("product_name", "").strip()
        category_name = row.get("category", "").strip()
        warehouse_name = row.get("warehouse_name", "").strip()
        qty_received_str = row.get("quantity_received", "").strip()
        qty_rejected_str = row.get("quantity_rejected", "").strip() or "0"
        purchase_price_str = row.get("purchase_price", "").strip()
        inventory_type = row.get("inventory_type", "").strip() or DEFAULT_INVENTORY_TYPE

        received_date = None
        if not received_date_str:
            errs.append("'received_date' is required")
        else:
            try:
                received_date = date.fromisoformat(received_date_str)
            except ValueError:
                errs.append(f"'received_date' must be YYYY-MM-DD, got '{received_date_str}'")

        supplier_id = suppliers.get(supplier_name)
        if not supplier_name:
            errs.append("'supplier_name' is required")
        elif supplier_id is None:
            errs.append(f"Supplier '{supplier_name}' not found — create it first")

        product = None
        if not product_name:
            errs.append("'product_name' is required")
        else:
            candidates = products_by_name.get(product_name, [])
            if not candidates:
                errs.append(f"Product '{product_name}' not found — create it first")
            elif supplier_id is not None:
                # Products carry a primary supplier_id — the product must belong to
                # the given supplier, mirroring the same rule enforced when
                # creating a receiving manually (see routers/receiving.py).
                matching = [p for p in candidates if p.supplier_id == supplier_id]
                if not matching:
                    errs.append(f"Product '{product_name}' is not assigned to supplier '{supplier_name}'")
                else:
                    product = matching[0]
            else:
                product = candidates[0]

        if product and category_name:
            actual_category = product.category.category_name if product.category else None
            if actual_category != category_name:
                errs.append(f"Product '{product_name}' belongs to category '{actual_category or 'none'}', not '{category_name}'")

        warehouse_id = None
        if warehouse_name:
            warehouse_id = warehouses.get(warehouse_name)
            if warehouse_id is None:
                errs.append(f"Warehouse '{warehouse_name}' not found")

        qty_received = None
        if not qty_received_str:
            errs.append("'quantity_received' is required")
        else:
            try:
                qty_received = int(float(qty_received_str))
                if qty_received <= 0:
                    errs.append("'quantity_received' must be greater than 0")
            except ValueError:
                errs.append(f"'quantity_received' must be a whole number, got '{qty_received_str}'")

        qty_rejected = 0
        try:
            qty_rejected = int(float(qty_rejected_str))
            if qty_rejected < 0:
                errs.append("'quantity_rejected' cannot be negative")
        except ValueError:
            errs.append(f"'quantity_rejected' must be a whole number, got '{qty_rejected_str}'")

        if qty_received is not None and qty_rejected > qty_received:
            errs.append(f"'quantity_rejected' ({qty_rejected}) cannot exceed 'quantity_received' ({qty_received})")

        purchase_price = None
        if not purchase_price_str:
            errs.append("'purchase_price' is required")
        else:
            try:
                purchase_price = float(purchase_price_str)
                if purchase_price < 0:
                    errs.append("'purchase_price' cannot be negative")
            except ValueError:
                errs.append(f"'purchase_price' must be a number, got '{purchase_price_str}'")

        if inventory_type not in INVENTORY_TYPES:
            errs.append(f"'inventory_type' must be one of: {', '.join(INVENTORY_TYPES)} — got '{inventory_type}'")

        if errs:
            errors.append({"row_number": i, "error_message": "; ".join(errs), "raw_data": json.dumps(row)})
        else:
            enriched = dict(row)
            enriched["_received_date"] = received_date
            enriched["_supplier_id"] = supplier_id
            enriched["_product_id"] = product.product_id
            enriched["_warehouse_id"] = warehouse_id
            enriched["_quantity_received"] = qty_received
            enriched["_quantity_rejected"] = qty_rejected
            enriched["_quantity_accepted"] = max(0, qty_received - qty_rejected)
            enriched["_purchase_price"] = purchase_price
            enriched["_inventory_type"] = inventory_type
            valid.append({"row": i, "data": enriched})
    return valid, errors


def _validate_bank_accounts(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    valid, errors = [], []
    required = ["bank_name", "account_number", "beneficiary_name"]
    for i, row in enumerate(rows, 1):
        errs = [f"'{c}' is required" for c in required if not row.get(c, "").strip()]
        if errs:
            errors.append({"row_number": i, "error_message": "; ".join(errs), "raw_data": json.dumps(row)})
        else:
            valid.append({"row": i, "data": row})
    return valid, errors


def _import_suppliers(valid_rows: list[dict], db: Session, username: str) -> tuple[int, list[dict]]:
    success, errors = 0, []
    for item in valid_rows:
        try:
            row = item["data"]
            obj = Supplier(
                supplier_name=row["supplier_name"].strip(),
                supplier_contact=row.get("supplier_contact", "").strip() or None,
                supplier_email=row.get("supplier_email", "").strip() or None,
                supplier_address=row.get("supplier_address", "").strip() or None,
                created_by=username,
            )
            db.add(obj)
            db.flush()
            success += 1
        except Exception as exc:
            errors.append({"row_number": item["row"], "error_message": str(exc), "raw_data": json.dumps(item["data"])})
    return success, errors


def _import_categories(valid_rows: list[dict], db: Session, username: str) -> tuple[int, list[dict]]:
    success, errors = 0, []
    for item in valid_rows:
        try:
            row = item["data"]
            obj = Category(
                category_name=row["category_name"].strip(),
                description=row.get("description", "").strip() or None,
                created_by=username,
            )
            db.add(obj)
            db.flush()
            success += 1
        except Exception as exc:
            errors.append({"row_number": item["row"], "error_message": str(exc), "raw_data": json.dumps(item["data"])})
    return success, errors


def _import_products(valid_rows: list[dict], db: Session, username: str) -> tuple[int, list[dict]]:
    success, errors = 0, []
    for item in valid_rows:
        try:
            row = item["data"]
            min_stock = 0.0
            try:
                min_stock = float(row.get("minimum_stock_level", 0) or 0)
            except (ValueError, TypeError):
                min_stock = 0.0
            obj = Product(
                product_name=row["product_name"].strip(),
                category_id=row.get("_category_id"),
                sale_price=float(row["sale_price"]),
                sku=row.get("sku", "").strip() or None,
                barcode=row.get("barcode", "").strip() or None,
                unit=row.get("unit", "PCS").strip() or "PCS",
                status=row.get("status", "Active").strip() or "Active",
                minimum_stock_level=min_stock,
                product_description=row.get("product_description", "").strip() or None,
                created_by=username,
            )
            db.add(obj)
            db.flush()
            success += 1
        except Exception as exc:
            errors.append({"row_number": item["row"], "error_message": str(exc), "raw_data": json.dumps(item["data"])})
    return success, errors


def _import_receivings(valid_rows: list[dict], db: Session, username: str) -> tuple[int, list[dict]]:
    """
    Mirrors routers/receiving.py::create_receiving — accepted qty updates
    inventory (with cost/bucket), and a Supplier Return auto-creates for any
    rejected qty. Kept as inline ORM logic (not a call into the router) to
    match this file's existing per-type import function pattern.
    """
    success, errors = 0, []
    for item in valid_rows:
        try:
            row = item["data"]
            receiving = Receiving(
                received_date=row["_received_date"],
                supplier_id=row["_supplier_id"],
                product_id=row["_product_id"],
                warehouse_id=row["_warehouse_id"],
                quantity_received=row["_quantity_received"],
                quantity_accepted=row["_quantity_accepted"],
                quantity_rejected=row["_quantity_rejected"],
                unit="PCS",
                purchase_price=row["_purchase_price"],
                inventory_type=row["_inventory_type"],
                created_by=username,
            )
            db.add(receiving)
            db.flush()

            if receiving.warehouse_id and receiving.quantity_accepted > 0:
                update_inventory_balance(
                    db,
                    product_id=receiving.product_id,
                    warehouse_id=receiving.warehouse_id,
                    qty_in=receiving.quantity_accepted,
                    qty_out=0,
                    transaction_type="RECEIVING",
                    reference_no=f"RCV-{receiving.receiving_id}",
                    inventory_type=receiving.inventory_type,
                    unit_cost_override=receiving.purchase_price,
                    created_by=username,
                )

            if receiving.quantity_rejected > 0 and receiving.supplier_id:
                db.add(SupplierReturn(
                    receiving_id=receiving.receiving_id,
                    supplier_id=receiving.supplier_id,
                    product_id=receiving.product_id,
                    warehouse_id=receiving.warehouse_id,
                    return_date=receiving.received_date,
                    quantity=receiving.quantity_rejected,
                    reason="Rejected at receiving",
                    status="Pending",
                    inventory_type=receiving.inventory_type,
                    created_by=username,
                ))

            success += 1
        except Exception as exc:
            errors.append({"row_number": item["row"], "error_message": str(exc), "raw_data": json.dumps(item["data"])})
    return success, errors


def _import_bank_accounts(valid_rows: list[dict], db: Session, username: str) -> tuple[int, list[dict]]:
    success, errors = 0, []
    for item in valid_rows:
        try:
            row = item["data"]
            is_active_str = str(row.get("is_active", "true")).strip().lower()
            is_active = is_active_str in ("1", "true", "yes", "active")
            obj = BankAccount(
                bank_name=row["bank_name"].strip(),
                account_number=row["account_number"].strip(),
                beneficiary_name=row["beneficiary_name"].strip(),
                is_active=is_active,
                created_by=username,
            )
            db.add(obj)
            db.flush()
            success += 1
        except Exception as exc:
            errors.append({"row_number": item["row"], "error_message": str(exc), "raw_data": json.dumps(item["data"])})
    return success, errors


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/templates/{import_type}")
def get_template(
    import_type: str,
    format: str = "csv",
    current_user: User = Depends(require_permission("bulk_upload.view")),
):
    """Return a CSV or XLSX template (columns + example row) for the given import type."""
    if import_type not in TEMPLATES:
        raise HTTPException(status_code=404, detail=f"Unknown import type. Supported: {', '.join(SUPPORTED_TYPES)}")
    if format not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="format must be 'csv' or 'xlsx'")
    tmpl = TEMPLATES[import_type]
    cols = tmpl["columns"]

    if format == "xlsx":
        try:
            from openpyxl import Workbook
        except ImportError:
            raise HTTPException(status_code=422, detail="XLSX templates require openpyxl to be installed on the server.")
        from fastapi.responses import Response
        wb = Workbook()
        ws = wb.active
        ws.title = import_type[:31]
        ws.append(cols)
        for example_row in tmpl["example"]:
            ws.append([example_row.get(c, "") for c in cols])
        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="template_{import_type}.xlsx"'},
        )

    lines = [",".join(cols)]
    for example_row in tmpl["example"]:
        lines.append(",".join(str(example_row.get(c, "")) for c in cols))
    csv_content = "\n".join(lines)
    from fastapi.responses import Response
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="template_{import_type}.csv"'},
    )


@router.post("/validate/{import_type}", response_model=BulkValidateResponse)
async def validate_file(
    import_type: str,
    file: UploadFile = File(...),
    current_user: User = Depends(require_permission("bulk_upload.view")),
    db: Session = Depends(get_db),
):
    """Parse and validate an uploaded file. Returns preview without importing."""
    if import_type not in SUPPORTED_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported import type. Supported: {', '.join(SUPPORTED_TYPES)}")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        rows = _parse_file(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not parse file: {exc}")

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in file")

    dispatch = {
        "suppliers": lambda r: _validate_suppliers(r),
        "categories": lambda r: _validate_categories(r),
        "products": lambda r: _validate_products(r, db),
        "bank_accounts": lambda r: _validate_bank_accounts(r),
        "receivings": lambda r: _validate_receivings(r, db),
    }
    valid_rows, error_list = dispatch[import_type](rows)

    preview = [item["data"] for item in valid_rows[:20]]  # first 20 valid rows for preview

    return BulkValidateResponse(
        import_type=import_type,
        total_rows=len(rows),
        valid_rows=len(valid_rows),
        invalid_rows=len(error_list),
        preview=preview,
        errors=[BulkImportErrorSchema(**e) for e in error_list],
    )


@router.post("/import/{import_type}", response_model=BulkImportResponse)
async def import_file(
    import_type: str,
    file: UploadFile = File(...),
    current_user: User = Depends(require_permission("bulk_upload.view")),
    db: Session = Depends(get_db),
):
    """Parse, validate, and import an uploaded file into the database."""
    if import_type not in SUPPORTED_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported import type. Supported: {', '.join(SUPPORTED_TYPES)}")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        rows = _parse_file(file.filename or "", content)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not parse file: {exc}")

    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in file")

    dispatch_validate = {
        "suppliers": lambda r: _validate_suppliers(r),
        "categories": lambda r: _validate_categories(r),
        "products": lambda r: _validate_products(r, db),
        "bank_accounts": lambda r: _validate_bank_accounts(r),
        "receivings": lambda r: _validate_receivings(r, db),
    }
    dispatch_import = {
        "suppliers": lambda v: _import_suppliers(v, db, current_user.username),
        "categories": lambda v: _import_categories(v, db, current_user.username),
        "products": lambda v: _import_products(v, db, current_user.username),
        "bank_accounts": lambda v: _import_bank_accounts(v, db, current_user.username),
        "receivings": lambda v: _import_receivings(v, db, current_user.username),
    }

    valid_rows, validation_errors = dispatch_validate[import_type](rows)
    success_count, import_errors = dispatch_import[import_type](valid_rows)

    all_errors = validation_errors + import_errors
    db.flush()

    # Persist import history
    history = BulkImportHistory(
        import_type=import_type,
        filename=file.filename,
        total_rows=len(rows),
        success_rows=success_count,
        error_rows=len(all_errors),
        status="completed",
        created_by=current_user.username,
    )
    db.add(history)
    db.flush()

    for err in all_errors:
        db.add(BulkImportError(
            import_id=history.import_id,
            row_number=err.get("row_number"),
            error_message=err["error_message"],
            raw_data=err.get("raw_data"),
        ))

    db.commit()

    return BulkImportResponse(
        import_id=history.import_id,
        import_type=import_type,
        total_rows=len(rows),
        success_rows=success_count,
        error_rows=len(all_errors),
        status="completed",
        errors=[BulkImportErrorSchema(**e) for e in all_errors],
    )


@router.get("/history", response_model=dict)
def import_history(
    import_type: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    current_user: User = Depends(require_permission("bulk_upload.view")),
    db: Session = Depends(get_db),
):
    q = db.query(BulkImportHistory)
    if import_type:
        q = q.filter(BulkImportHistory.import_type == import_type)
    q = q.order_by(BulkImportHistory.created_at.desc())
    total = q.count()
    items = q.offset((page - 1) * limit).limit(limit).all()
    return {
        "total": total, "page": page, "limit": limit,
        "items": [
            {
                "import_id": h.import_id,
                "import_type": h.import_type,
                "filename": h.filename,
                "total_rows": h.total_rows,
                "success_rows": h.success_rows,
                "error_rows": h.error_rows,
                "status": h.status,
                "created_by": h.created_by,
                "created_at": h.created_at.isoformat() if h.created_at else None,
            }
            for h in items
        ],
    }
